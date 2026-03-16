"""
Endpoint de exportación de datos para consultores del BID.
Soporta CSV y XLSX con filtros por departamento, sector, etc.
"""
import csv
import io
import tempfile
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, cast, Date
from typing import Optional
from datetime import date

from db.session import get_db
from db.models import Vacante

router = APIRouter(prefix="/vacantes", tags=["exportar"])

# Campos a exportar
EXPORT_COLUMNS = [
    "codigo_vacante", "titulo_vacante", "cargo", "nivel_estudios",
    "rango_salarial", "tipo_contrato", "cantidad_vacantes",
    "meses_experiencia_cargo", "sector_economico", "departamento",
    "municipio", "fecha_publicacion", "fecha_vencimiento",
    "teletrabajo", "discapacidad", "plaza_practica", "teletrabajo"
]

MAX_EXPORT_ROWS = 50000


def _build_filtered_query(
    departamento: Optional[str],
    sector: Optional[str],
    tipo_contrato: Optional[str],
    nivel_estudios: Optional[str],
    fecha_desde: Optional[date],
    fecha_hasta: Optional[date],
):
    stmt = select(Vacante)
    if departamento:
        stmt = stmt.where(Vacante.departamento == departamento)
    if sector:
        stmt = stmt.where(Vacante.sector_economico == sector)
    if tipo_contrato:
        stmt = stmt.where(Vacante.tipo_contrato == tipo_contrato)
    if nivel_estudios:
        stmt = stmt.where(Vacante.nivel_estudios == nivel_estudios)
    if fecha_desde:
        stmt = stmt.where(cast(Vacante.fecha_publicacion, Date) >= fecha_desde)
    if fecha_hasta:
        stmt = stmt.where(cast(Vacante.fecha_publicacion, Date) <= fecha_hasta)
    return stmt.order_by(Vacante.fecha_publicacion.desc()).limit(MAX_EXPORT_ROWS)


def _vacante_to_row(v) -> dict:
    return {
        "codigo_vacante": v.codigo_vacante,
        "titulo_vacante": v.titulo_vacante,
        "cargo": v.cargo,
        "nivel_estudios": v.nivel_estudios,
        "rango_salarial": v.rango_salarial,
        "tipo_contrato": v.tipo_contrato,
        "cantidad_vacantes": v.cantidad_vacantes,
        "meses_experiencia_cargo": v.meses_experiencia_cargo,
        "sector_economico": v.sector_economico,
        "departamento": v.departamento,
        "municipio": v.municipio,
        "fecha_publicacion": str(v.fecha_publicacion) if v.fecha_publicacion else "",
        "fecha_vencimiento": str(v.fecha_vencimiento) if v.fecha_vencimiento else "",
        "teletrabajo": v.teletrabajo,
        "discapacidad": v.discapacidad,
        "plaza_practica": v.plaza_practica,
    }


@router.get("/exportar")
async def export_vacantes(
    formato: str = Query("csv", regex="^(csv|xlsx)$"),
    departamento: Optional[str] = None,
    sector: Optional[str] = None,
    tipo_contrato: Optional[str] = None,
    nivel_estudios: Optional[str] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    db: AsyncSession = Depends(get_db),
):
    stmt = _build_filtered_query(departamento, sector, tipo_contrato, nivel_estudios, fecha_desde, fecha_hasta)
    result = await db.execute(stmt)
    vacantes = result.scalars().all()

    rows = [_vacante_to_row(v) for v in vacantes]

    if formato == "csv":
        return _generate_csv_response(rows)
    else:
        return _generate_xlsx_response(rows)


def _generate_csv_response(rows: list[dict]):
    output = io.StringIO()
    if rows:
        writer = csv.DictWriter(output, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    else:
        writer = csv.writer(output)
        writer.writerow(EXPORT_COLUMNS)

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=vacantes_colombia.csv"},
    )


def _generate_xlsx_response(rows: list[dict]):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Vacantes Colombia"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")

    headers = list(rows[0].keys()) if rows else EXPORT_COLUMNS
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()

    return FileResponse(
        tmp.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="vacantes_colombia.xlsx",
    )
